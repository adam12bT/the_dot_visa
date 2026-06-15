"""
ExcelGenerator — reproduces the exact layout, colours and formulas
of Liste_des_candidatures_Vivatech_2026.xlsx from a scored DataFrame.
"""
from __future__ import annotations
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Exact colours (ARGB) ──────────────────────────────────────────────────────
YELLOW    = "FFFFE599"
RED       = "FFFF0000"
ORANGE    = "FFFF9900"
GREY999   = "FF999999"
WHITE     = "FFFFFFFF"
LGREY     = "FFF8F9FA"
GREEN_BG  = "FFB7E1CD"
GREEN_SA  = "FFB6D7A8"
PURPLE    = "FFFF00FF"
BLUE      = "FF4A86E8"
AMBER     = "FFFF9900"
GREEN_SC  = "FF00FF00"

# Sheet 1 palette (matches the purple-header reference)
HEADER_BG   = "FF4A3C8C"
HEADER_FG   = "FFFFFFFF"
ZEBRA_LIGHT = "FFFFFFFF"
ZEBRA_DARK  = "FFF4F4F7"


def _fill(rgb: str) -> PatternFill:
    return PatternFill("solid", fgColor=rgb)


def _font(bold=False, size=None, color="FF000000") -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color)


def _c(ws, row, col, value=None, fg=None, bold=False, size=None,
       halign="left", valign="center", wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, size=size, color="FF000000")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if fg:
        cell.fill = _fill(fg)
    return cell


# ── delegation detection ──────────────────────────────────────────────────────
OFFICIAL_DELEGATIONS = [
    "expertise france", "fondation tunisie pour le développement",
    "fondation tunisie", "giz tunisie", "giz", "cepex", "mtc",
]


class ExcelGenerator:

    @staticmethod
    def generate(df_scored: pd.DataFrame) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)
        ExcelGenerator._build_reponses(wb, df_scored)
        ExcelGenerator._build_methode(wb)
        ExcelGenerator._build_evaluation(wb, df_scored)
        ExcelGenerator._build_classement(wb, df_scored)
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    # ── Sheet 1 : Réponses au formulaire 1 ────────────────────────────────────
    @staticmethod
    def _build_reponses(wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Réponses au formulaire 1")

        headers = [
            "Horodateur", "Adresse e-mail", "Nom de la start-up ",
            "Votre nom et prénom", "Votre e-mail", "Numéro de téléphone",
            "Votre position dans la start-up",
            "Quel est le nom de votre start-up?",
            "Quel est le secteur d'activité de votre start-up ?",
            "Le lien du site web ( ou page LinkedIn principale) ",
            "De quel programme de The Dot votre start-up a-t-elle bénéficié ? ",
            "Votre start-up a-t-elle le label Start-up Act ? ",
            "Quel âge a votre start-up?",
            "Quelle est votre région d'implantation?",
            "Décrivez brièvement votre produit/solution (maximum 100 mots)",
            "Nombre d'employés à plein temps au sein de l'équipe  ",
            "Lien vers la version la plus récente de votre Pitch Deck ( YouTube, drive...)",
            "Votre start-up génère-t-elle un chiffre d\u2019affaires ?",
            "Quel est le chiffre d'affaire que votre startup a généré en 2025 ? ",
            "Traction commerciale sur le marché local ",
            "Votre start-up est-elle déjà présente sur le marché européen ou international ? ",
            "Si oui, précisez brièvement (Pays cibles, clients majeurs ou CA à l'export)  ",
            "Si oui, merci de préciser vos principaux clients ou partenaires à l\u2019international.\n",
            "Si oui, merci de préciser le chiffre d\u2019affaires estimé ou le nombre de contrats signés à l\u2019étranger",
            "Avez-vous déjà levé des fonds (equity ou quasi-equity)?",
            "Si oui, quel est le montant total cumulé et l'année de la dernière levée ? ( précisez la devise)",
            "Quels sont les objectifs prioritaires de votre participation à VivaTech ? (Maximum 2 choix) ",
            "Si oui, précisez le nombre d'utilisateurs ayant validé votre solution ",
            "Pouvez-vous préciser l\u2019état d\u2019avancement de votre projet ainsi que votre calendrier de mise sur le marché (go-to-market) ? ",
            "Avez-vous déjà levé des fonds (equity ou quasi equity)?",
            "Si oui, quel est le montant du ticket et l\u2019année d\u2019obtention?",
            "Fonction de cette personne au sein de la start-up",
            "Avez-vous déjà participé à un salon international auparavant?",
            "Si Oui, Le(s)quel(s) ?",
            "Nom & prénom de la personne qui représentera physiquement la start-up à Paris ",
            "La fonction au sein de l'entreprise de la personne participante",
            "Adresse mail ",
            "Contact téléphone ",
            "La personne représentant la structure dispose-t-elle d\u2019un passeport valide au moins 3 mois à compter de son entrée en France (soit à compter du 15 juin 2026)  ",
            " Y a-t-il une femme parmi les co-fondateurs ?  ",
            "La personne représentant la structure dispose-t-elle d\u2019un Visa Schengen valide pour la période du 17 au 20 Juin 2026? ",
            "Si oui, quel a été votre chiffre d'affaires global en 2025 (en TND)?",
            "Quel est le stade de développement actuel de votre produit?",
            "Si oui, à quel salon (nom et année) ? Faisiez-vous partie d'une délégation officielle soutenue par Expertise France, la Fondation Tunisie pour le Développement ou la GIZ Tunisie ?",
            "Évaluation Post-Événement  ",
            "Autorisation d'utilisation de l'image (Droit à l'image)  ",
            "Colonne 44", "Colonne 42", "Colonne 41",
        ]

        col_map = {
            0:  "Horodateur",         1:  "Email formulaire",
            2:  "Startup",            3:  "Contact (nom)",
            4:  "Email contact",      5:  "Téléphone contact",
            6:  "Position dans startup", 7: "Startup",
            8:  "Secteur",            9:  "Site web / LinkedIn",
            10: "Programme The Dot",  11: "Label Startup Act (info)",
            12: "Âge startup",        13: "Région",
            14: "Description produit",15: "Nb employés",
            16: "Pitch Deck",         17: "Génère un CA",
            18: "CA 2025 (déclaré)",  19: "Traction locale",
            20: "Présence internationale", 21: "Pays / marchés ciblés",
            22: "Clients / partenaires intl",
            23: "CA ou contrats à l'étranger",
            24: "Levée fonds (equity)",   25: "Montant levée (equity)",
            26: "Objectifs VivaTech (qualitatif)", 27: "Utilisateurs pilotes",
            28: "État avancement / GTM",
            29: "Levée fonds (quasi-equity)", 30: "Montant levée (quasi-equity)",
            31: "Détail objectifs",   32: "Salon intl passé (info)",
            33: "Salon(s) précédent(s)", 34: "Représentant (nom)",
            35: "Représentant (fonction)", 36: "Représentant (email)",
            37: "Représentant (tél)", 38: "Passport valide",
            39: "Femme co-fondatrice",40: "Visa Schengen valide",
            41: "CA 2025 (TND)",      42: "Stade produit",
            43: "Salon + délégation officielle",
            44: "Engagement éval. post-event", 45: "Droit à l'image",
        }

        # ── Header row 1: purple bar, white bold text ──
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
            cell.fill = _fill(HEADER_BG)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 32

        # Colour map for the "name" cell only (column C)
        COLOR_MAP = {
            "yellow": YELLOW, "jaune":  YELLOW,
            "red":    RED,    "rouge":  RED,
            "orange": ORANGE,
            "grey":   GREY999, "gray":  GREY999, "gris": GREY999,
        }

        NAME_COL_IDX = 3  # column C

        for r_idx, (_, row) in enumerate(df.iterrows()):
            er = r_idx + 2
            zebra_fg = ZEBRA_LIGHT if r_idx % 2 == 0 else ZEBRA_DARK

            # Resolve accent colour for the startup-name cell
            accent = None
            tag = row.get("Couleur ligne", None)
            if isinstance(tag, str) and tag.strip():
                accent = COLOR_MAP.get(tag.strip().lower())
            if accent is None:
                ddl_ok = row.get("✅ Respect DDL", None)
                dot    = row.get("✅ Bénéficiaire The Dot", None)
                vis    = row.get("✅ Visa valide", None)
                if ddl_ok is False:
                    accent = ORANGE
                elif dot is False:
                    accent = GREY999
                elif dot is True and vis is False:
                    accent = RED

            for ci in range(1, len(headers) + 1):
                df_col = col_map.get(ci - 1)
                val = None
                if df_col and df_col in row.index:
                    raw = row[df_col]
                    val = None if pd.isna(raw) else (raw if ci == 1 else str(raw))

                cell = ws.cell(row=er, column=ci, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                if ci == NAME_COL_IDX and accent:
                    white_text = accent in (RED, ORANGE)
                    cell.font = Font(
                        name="Arial", bold=True, size=11,
                        color=HEADER_FG if white_text else "FF000000",
                    )
                    cell.fill = _fill(accent)
                else:
                    cell.font = _font(size=11)
                    cell.fill = _fill(zebra_fg)



        
        # Column widths matching original
# Auto-size columns based on content

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)

            max_length = 0

            # Check header
            header_length = len(str(headers[col_num - 1]))
            max_length = max(max_length, header_length)

            # Check data rows
            for row_num in range(2, ws.max_row + 1):
                value = ws.cell(row=row_num, column=col_num).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))

            # Add padding and limit maximum width
            ws.column_dimensions[column_letter].width = min(max_length + 5, 60)











        for r in range(2, len(df) + 2):
            ws.row_dimensions[r].height = 22.5

        # AutoFilter dropdowns on the header row
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(df) + 1}"

        # Legend
        legend_row = len(df) + 4
        _c(ws, legend_row,     1, "Code couleurs", bold=True)
        _c(ws, legend_row,     3, "Candidatures soumises après la ddl")
        _c(ws, legend_row + 1, 3, "N'ont pas de visa ")
        _c(ws, legend_row + 2, 3, "Ne font pas partie de The Dot ")
        ws.cell(legend_row,     2).fill = _fill(ORANGE)
        ws.cell(legend_row + 1, 2).fill = _fill(RED)
        ws.cell(legend_row + 2, 2).fill = _fill(GREY999)

        ws.freeze_panes = "D2"

    # ── Sheet 2 : Methode de scoring ──────────────────────────────────────────
    @staticmethod
    def _build_methode(wb: Workbook) -> None:
        ws = wb.create_sheet("Methode de scoring ")

        _c(ws, 1, 2, "Critères d'éligibilité ", fg=GREEN_SA, bold=True, halign="center")
        ws.merge_cells("B1:F2")

        _c(ws, 3, 2, "Bénéficiaires de The Dot ", fg=GREEN_SA)
        _c(ws, 3, 3, "Non Selectionné par une autre délégation", fg=GREEN_SA)
        _c(ws, 3, 4, "Respect de la DDL", fg=GREEN_SA, halign="center")
        _c(ws, 3, 5, "Passport Valide", fg=GREEN_SA, halign="center")
        _c(ws, 3, 6, "Visa Valide", fg=GREEN_SA, halign="center")

        _c(ws, 5, 2, "Critères ", bold=True)
        _c(ws, 5, 3, "Sous-critères ", bold=True)
        _c(ws, 5, 4, "Réponses ", bold=True)
        _c(ws, 5, 5, "Notation", bold=True)

        _c(ws, 6, 1, "Critères d'évaluation ", bold=True)
        ws.merge_cells("A6:A27")

        # Inclusivité
        _c(ws, 6, 2, "Inclusivité ", fg=PURPLE, bold=True); ws.merge_cells("B6:B11")
        _c(ws, 6, 3, "Région ");  ws.merge_cells("C6:C7")
        _c(ws, 6, 4, "Grand Tunis ");   _c(ws, 6, 5, 0.0)
        _c(ws, 7, 4, "Hors Tunis");     _c(ws, 7, 5, 1.0)
        _c(ws, 8, 3, "Genre");   ws.merge_cells("C8:C9")
        _c(ws, 8, 4, "Parmi les co-fondateur, il n' y'a pas de femme ");  _c(ws, 8, 5, 0.0)
        _c(ws, 9, 4, "Parmi les co-fondateur, exsite une femme ");         _c(ws, 9, 5, 1.0)
        _c(ws, 10, 3, "Label startup ");  ws.merge_cells("C10:C11")
        _c(ws, 10, 4, "Pas de label ");   _c(ws, 10, 5, 0.0)
        _c(ws, 11, 4, "Label octroyé");   _c(ws, 11, 5, 1.0)

        # Maturité
        _c(ws, 12, 2, "Maturité du projet ", fg=BLUE, bold=True); ws.merge_cells("B12:B21")
        _c(ws, 12, 3, "Age ");  ws.merge_cells("C12:C13")
        _c(ws, 12, 4, "Moins de 2 ans ");  _c(ws, 12, 5, 0.0)
        _c(ws, 13, 4, "Plus que 2 ans ");  _c(ws, 13, 5, 1.0)
        _c(ws, 14, 3, "Nombre d'employés ");  ws.merge_cells("C14:C17")
        _c(ws, 14, 4, "Moins de 2");    _c(ws, 14, 5, 0.0)
        _c(ws, 15, 4, "Entre 2 et 4");  _c(ws, 15, 5, 1.0)
        _c(ws, 16, 4, "Entre 5 et 9");  _c(ws, 16, 5, 2.0)
        _c(ws, 17, 4, "10 et plus ");   _c(ws, 17, 5, 3.0)
        _c(ws, 18, 3, "Chiffres d'affaires ");  ws.merge_cells("C18:C19")
        _c(ws, 18, 4, "Pas Exsitant");  _c(ws, 18, 5, 0.0)
        _c(ws, 19, 4, "Existant ");     _c(ws, 19, 5, 1.0)
        _c(ws, 20, 3, "Levée de fonds ");  ws.merge_cells("C20:C21")
        _c(ws, 20, 4, "Aucune levée réalisée ");  _c(ws, 20, 5, 0.0)
        _c(ws, 21, 4, "Ayant levé");             _c(ws, 21, 5, 1.0)

        # Pertinence
        _c(ws, 22, 2, "Pertinence de la participation au salon ", fg=AMBER, bold=True)
        ws.merge_cells("B22:B28")
        _c(ws, 22, 3, "Participation à un salon ");  ws.merge_cells("C22:C23")
        _c(ws, 22, 4, "N'ayant pas dejà participé ");  _c(ws, 22, 5, 0.0)
        _c(ws, 23, 4, "Ayant déja participé");          _c(ws, 23, 5, 1.0)
        _c(ws, 24, 3, "Objectifs ")
        _c(ws, 24, 4, "App qualitative");  ws.merge_cells("D24:E24")
        _c(ws, 25, 3, "Votre startup est-elle déjà présente sur des marchés internationaux ?")
        ws.merge_cells("C25:C26")
        _c(ws, 25, 4, "Non ");   _c(ws, 25, 5, 0.0)
        _c(ws, 26, 4, "Oui ");   _c(ws, 26, 5, 1.0)
        _c(ws, 27, 3, "Si oui, merci de préciser les pays et/ou les marchés concernés")
        ws.merge_cells("C27:C28")
        _c(ws, 27, 4, "Marché européen");  _c(ws, 27, 5, 1.0)
        _c(ws, 28, 4, "Autre marché ");    _c(ws, 28, 5, 0.0)

    # ── Sheet 3 : Evaluation ──────────────────────────────────────────────────
    @staticmethod
    def _build_evaluation(wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Evaluation")

        for col, val, fg, halign in [
            (1,  "Nom de la startup ",    None,     "left"),
            (2,  "Description ",          None,     "left"),
            (3,  "Adresse mail ",         None,     "left"),
            (4,  "Bénéficiaires de The Dot ", None, "left"),
            (5,  "Respect de la DDL",     None,     "center"),
            (6,  "Passport Valide",       None,     "center"),
            (7,  "Visa Valide",           None,     "center"),
            (8,  "Non Selectionné par une autre deleg", None, "left"),
            (9,  "Inclusivité ",          PURPLE,   "center"),
            (12, "Maturité du projet ",   BLUE,     "center"),
            (16, "Pertinence de la participation au salon ", AMBER, "center"),
            (20, "Score ",                GREEN_SC, "left"),
        ]:
            _c(ws, 1, col, val, fg=fg, bold=True, size=12, halign=halign)

        ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2"); ws.merge_cells("C1:C2")
        ws.merge_cells("D1:D2"); ws.merge_cells("E1:E2"); ws.merge_cells("F1:F2")
        ws.merge_cells("G1:G2"); ws.merge_cells("H1:H2")
        ws.merge_cells("I1:K1"); ws.merge_cells("L1:O1"); ws.merge_cells("P1:S1")
        ws.merge_cells("T1:T2")

        for col, val, fg in [
            (9,  "Région ",                PURPLE),
            (10, "Genre ",                 PURPLE),
            (11, "Label startup act",      PURPLE),
            (12, "Age ",                   BLUE),
            (13, "Nombre d'employés ",     BLUE),
            (14, "Chiffe d'affaire ",      BLUE),
            (15, "Levée de fonds",         BLUE),
            (16, "Participation à un salon auparavant ", AMBER),
            (17, "Objectifs de participation ", AMBER),
            (18, "Votre startup est-elle déjà présente sur des marchés internationaux ?", AMBER),
            (19, "Si oui, merci de préciser les pays et/ou les marchés concernés", AMBER),
        ]:
            _c(ws, 2, col, val, fg=fg, bold=True, size=12)

        ws.column_dimensions["C"].width = 38.38

        eligible_count = 0
        for r_offset, (_, row) in enumerate(df.iterrows()):
            er = r_offset + 3

            dot  = bool(row.get("✅ Bénéficiaire The Dot", False))
            pas  = bool(row.get("✅ Passport valide", False))
            oth  = bool(row.get("✅ Non sélec. autre déleg.", False))
            vis  = bool(row.get("✅ Visa valide", False))
            elig = dot and pas and oth

            if elig:
                row_fg = WHITE if eligible_count % 2 == 0 else LGREY
                eligible_count += 1
            else:
                row_fg = RED

            _c(ws, er, 1, str(row.get("Startup", "") or ""), fg=row_fg)
            _c(ws, er, 2, str(row.get("Contact (nom)", "") or ""), fg=row_fg)
            _c(ws, er, 3, str(row.get("Email contact", "") or ""), fg=row_fg)

            deleg_val = not bool(row.get("✅ Non sélec. autre déleg.", True))
            for col_i, bval in {4: dot, 5: True, 6: pas, 7: vis, 8: deleg_val}.items():
                c = ws.cell(row=er, column=col_i, value=bval)
                c.font = _font()
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.fill = _fill(GREEN_BG)

            if elig:
                s_reg  = float(row.get("S: Région hors Gd Tunis (0-1)", 0))
                s_gen  = float(row.get("S: Femme co-fond. (0-1)", 0))
                s_lbl  = 1.0 if str(row.get("Label Startup Act (info)", "")).lower() == "oui" else 0.0
                s_age  = float(row.get("S: Âge ≥2 ans (0-1)", 0))
                s_emp  = float(row.get("S: Nb employés (0-3)", 0))
                s_ca   = float(row.get("S: CA existant (0-1)", 0))
                s_fond = float(row.get("S: Levée fonds (0-1)", 0))
                s_sal  = 1.0 if str(row.get("Salon intl passé (info)", "")).lower() == "oui" else 0.0
                obj    = str(row.get("Objectifs VivaTech (qualitatif)", "") or "")
                s_intl = float(row.get("S: Présence intl (0-1)", 0))
                s_eu   = float(row.get("S: Marché européen (0-1)", 0))

                for col_i, val in [
                    (9, s_reg), (10, s_gen), (11, s_lbl),
                    (12, s_age), (13, s_emp), (14, s_ca), (15, s_fond),
                    (16, s_sal),
                ]:
                    c = ws.cell(row=er, column=col_i, value=val)
                    c.font = _font()
                    c.alignment = Alignment(horizontal="right", vertical="center")

                _c(ws, er, 17, obj, fg=row_fg, halign="left")

                for col_i, val in [(18, s_intl), (19, s_eu)]:
                    c = ws.cell(row=er, column=col_i, value=val)
                    c.font = _font()
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.fill = _fill(row_fg)

                i = get_column_letter(9);  j = get_column_letter(10)
                l = get_column_letter(12); m = get_column_letter(13)
                n = get_column_letter(14); o = get_column_letter(15)
                r_ = get_column_letter(18); s = get_column_letter(19)
                formula = f"={i}{er}+{j}{er}+{l}{er}+{m}{er}+{n}{er}+{o}{er}+{r_}{er}+{s}{er}"
                c = ws.cell(row=er, column=20, value=formula)
                c.font = _font()
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                for col_i in range(9, 20):
                    c = ws.cell(row=er, column=col_i)
                    c.fill = _fill(RED)

    # ── Sheet 4 : Classement ──────────────────────────────────────────────────
    @staticmethod
    def _build_classement(wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Classement")

        for ci, val in enumerate(["Nom de la startup ", "Score ", "Genre "], 1):
            _c(ws, 1, ci, val, bold=True, size=12)
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:C2")

        eligible = df[
            df["✅ Bénéficiaire The Dot"] &
            df["✅ Passport valide"] &
            df["✅ Non sélec. autre déleg."]
        ].copy().sort_values("SCORE TOTAL /10", ascending=False)

        for r_offset, (_, row) in enumerate(eligible.iterrows()):
            er = r_offset + 3
            score = float(row.get("SCORE TOTAL /10", 0))
            genre = "F" if int(row.get("S: Femme co-fond. (0-1)", 0)) == 1 else "H"

            _c(ws, er, 1, str(row.get("Startup", "")), fg=YELLOW)
            c_score = ws.cell(row=er, column=2, value=score)
            c_score.font = _font(); c_score.fill = _fill(YELLOW)
            c_score.alignment = Alignment(horizontal="right", vertical="center")
            c_genre = ws.cell(row=er, column=3, value=genre)
            c_genre.font = _font(); c_genre.fill = _fill(YELLOW)
            c_genre.alignment = Alignment(horizontal="right", vertical="center")
